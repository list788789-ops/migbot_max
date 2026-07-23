"""
purchases/parser.py — разбор Excel-накладных поставщиков в строки документа
«Поступление товаров и услуг».

Чистый модуль без БД и без FastAPI: на входе — байты файла, на выходе — dataclass'ы.
Так его можно звать и из роутов (webforms.py), и из фоновой задачи, и из тестов.
Работа с БД (сопоставление номенклатуры, создание документа) — в matching.py и service.py.

Ключевое решение: номер строки заголовков НЕ хардкодится. У каждого поставщика своя
шапка (реквизиты, «Грузополучатель», объединённые ячейки), и таблица начинается на
разной строке. find_header ищет строку, где сошлись минимум MIN_HEADER_HITS известных
названий колонок. Если не нашли — не гадаем, а поднимаем HeaderNotFound: пусть человек
один раз укажет колонки руками, сохраним это в профиль поставщика и больше не спросим.

Профиль поставщика (SupplierProfile) — то, что делает загрузку повторяемой. Хранится
в БД по ИНН, здесь только структура и применение. Профиль перекрывает автопоиск:
если он задан, эвристика не запускается вообще.

Деньги — только Decimal. float на суммах даёт расхождение с итогом накладной на копейки,
а копейки в первичке — это расхождение с поставщиком и вопрос на сверке.

Хэш файла (sha256 по байтам) — защита от повторной загрузки одной и той же накладной.
Уникальный индекс по нему в таблице документов; проверять ДО разбора, разбор недешёвый.

Контрольная сумма: если в файле нашёлся итог, сверяем с суммой строк. Расхождение не
роняет разбор — оно попадает в ParseResult.warnings, документ создаётся, но помечается
как требующий проверки. Молча проглатывать расхождение нельзя, но и отказываться грузить
из-за копейки округления — тоже.

2026-07: единицы измерения приводятся к каноническому виду (шт/кг/т/м/м2/м3/л/компл).
Поставщики пишут «шт.», «ШТ», «штук» — без нормализации сопоставление номенклатуры
даёт ложные расхождения по упаковке.
"""

from __future__ import annotations

import hashlib
import io
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

log = logging.getLogger("purchases.parser")


# --- Ошибки -----------------------------------------------------------------

class ParseError(Exception):
    """Базовая: файл не удалось разобрать."""


class HeaderNotFound(ParseError):
    """Строка заголовков не найдена и профиль поставщика не задан."""


# --- Словарь колонок --------------------------------------------------------
# Ключ — внутреннее имя поля, значения — начала названий колонок у поставщиков
# (сравнение по startswith после нормализации, поэтому «Кол-во, шт» ловится «кол-во»).

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "article": ("артикул", "код", "код товара", "кат. номер", "номенклатурный номер"),
    "name":    ("наименование", "товар", "номенклатура", "материал", "описание"),
    "unit":    ("ед", "единица", "ед.изм"),
    "qty":     ("кол-во", "количество", "колич"),
    "price":   ("цена", "цена без ндс", "цена за ед"),
    "amount":  ("сумма", "стоимость", "сумма без ндс"),
    "vat_rate": ("ставка ндс", "% ндс", "ндс, %", "ставка"),
    "vat_sum": ("сумма ндс", "ндс сумма", "в т.ч. ндс"),
    "total":   ("всего с ндс", "сумма с ндс", "всего"),
}

# Обязательный минимум, чтобы признать строку шапкой таблицы.
REQUIRED_FIELDS = ("name", "qty")
MIN_HEADER_HITS = 3
MAX_HEADER_SCAN = 40  # глубже 40 строк шапки не бывает даже у самых многословных

# Строки, которые выглядят как позиции, но ими не являются.
STOP_ROW_RE = re.compile(
    r"^\s*(итого|всего|в\s*т\.?\s*ч|сумма\s+прописью|руководитель|отпустил|получил|"
    r"главный\s+бухгалтер|м\.?п\.?)",
    re.IGNORECASE,
)

UNIT_CANON: dict[str, str] = {
    "шт": "шт", "штук": "шт", "штука": "шт", "шт.": "шт", "pcs": "шт",
    "кг": "кг", "килограмм": "кг",
    "т": "т", "тн": "т", "тонна": "т", "тонн": "т",
    "м": "м", "пм": "м", "п.м": "м", "п/м": "м", "погм": "м", "метр": "м",
    "м2": "м2", "кв.м": "м2", "квм": "м2", "м²": "м2",
    "м3": "м3", "куб.м": "м3", "кубм": "м3", "м³": "м3",
    "л": "л", "литр": "л",
    "компл": "компл", "комплект": "компл", "к-т": "компл", "набор": "компл",
    "уп": "уп", "упак": "уп", "упаковка": "уп",
}


# --- Структуры данных -------------------------------------------------------

@dataclass
class SupplierProfile:
    """Сохранённая раскладка файла конкретного поставщика.

    Заполняется один раз человеком (или автопоиском, если он сработал) и дальше
    применяется без эвристики. inn — ключ хранения в БД.
    """
    inn: str | None = None
    sheet: str | None = None          # None → активный лист
    header_row: int | None = None     # 1-based, как в Excel
    columns: dict[str, int] = field(default_factory=dict)  # поле → индекс колонки (0-based)
    skip_rows_after_header: int = 0   # у некоторых поставщиков под шапкой строка нумерации


@dataclass
class InvoiceLine:
    """Одна позиция накладной — как она пришла в файле, без сопоставления."""
    row: int                    # номер строки в Excel, для показа человеку при ошибке
    name: str
    qty: Decimal
    unit: str                   # канонический
    unit_raw: str               # как было в файле
    article: str | None = None
    price: Decimal | None = None
    amount: Decimal | None = None
    vat_rate: str | None = None

    @property
    def computed_amount(self) -> Decimal | None:
        if self.price is None:
            return None
        return (self.qty * self.price).quantize(Decimal("0.01"))


@dataclass
class ParseResult:
    lines: list[InvoiceLine]
    file_hash: str
    header_row: int
    columns: dict[str, int]
    sheet: str
    declared_amount: Decimal | None = None  # итог без НДС — с ним сверяем строки
    declared_total: Decimal | None = None   # итог с НДС — только для показа
    warnings: list[str] = field(default_factory=list)

    @property
    def lines_total(self) -> Decimal:
        return sum(
            (ln.amount if ln.amount is not None else (ln.computed_amount or Decimal(0))
             for ln in self.lines),
            Decimal(0),
        )

    @property
    def needs_review(self) -> bool:
        return bool(self.warnings)


# --- Утилиты ----------------------------------------------------------------

def file_hash(data: bytes) -> str:
    """sha256 по байтам файла. Проверять ДО разбора — дешевле, чем парсить дубль."""
    return hashlib.sha256(data).hexdigest()


def _norm(v) -> str:
    """Нормализация текста ячейки для сравнения с алиасами."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", s)


def to_decimal(v) -> Decimal | None:
    """Число из ячейки. Терпит '1 234,56', '1234.56', неразрывные пробелы, None.

    Excel обычно отдаёт float — приводим через str, чтобы не тащить погрешность
    двоичного представления в деньги.
    """
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def canon_unit(raw) -> tuple[str, str]:
    """('шт.', ) → ('шт', 'шт.'). Возвращает (канон, как_было)."""
    raw_s = str(raw or "").strip()
    key = _norm(raw_s).rstrip(".")
    return UNIT_CANON.get(key, key or ""), raw_s


# --- Поиск шапки ------------------------------------------------------------

def find_header(ws, max_scan: int = MAX_HEADER_SCAN) -> tuple[int, dict[str, int]]:
    """Ищет строку заголовков таблицы. Возвращает (номер строки 1-based, {поле: индекс}).

    Берём первую строку, где сошлись MIN_HEADER_HITS полей И присутствуют REQUIRED_FIELDS.
    Порядок алиасов внутри поля важен: длинные варианты проверяются раньше коротких,
    иначе «сумма ндс» уедет в «amount» по алиасу «сумма».
    """
    best: tuple[int, dict[str, int]] | None = None

    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        mapping: dict[str, int] = {}
        for cell in row:
            text = _norm(cell.value)
            if not text:
                continue
            match_field, match_len = None, -1
            for fld, aliases in HEADER_ALIASES.items():
                for alias in aliases:
                    # самое длинное совпадение выигрывает: «сумма ндс» > «сумма»
                    if text.startswith(alias) and len(alias) > match_len:
                        match_field, match_len = fld, len(alias)
            if match_field and match_field not in mapping:
                mapping[match_field] = cell.column - 1

        if len(mapping) >= MIN_HEADER_HITS and all(f in mapping for f in REQUIRED_FIELDS):
            best = (row[0].row, mapping)
            break

    if best is None:
        raise HeaderNotFound(
            "Не удалось определить строку заголовков. "
            "Укажите колонки вручную — раскладка сохранится в профиль поставщика."
        )
    return best


# --- Разбор -----------------------------------------------------------------

def parse_invoice(data: bytes, profile: SupplierProfile | None = None) -> ParseResult:
    """Главная точка входа. data — байты xlsx, profile — сохранённая раскладка или None.

    read_only=False сознательно: в read_only недоступны ws.merged_cells, а у половины
    поставщиков шапка с объединениями. Файлы накладных небольшие, память не проблема.
    """
    fh = file_hash(data)
    warnings: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:  # noqa: BLE001 — важно отдать понятную ошибку наверх
        raise ParseError(f"Файл не читается как xlsx: {e}") from e

    if profile and profile.sheet and profile.sheet in wb.sheetnames:
        ws = wb[profile.sheet]
    else:
        ws = wb.active

    if profile and profile.header_row and profile.columns:
        header_row, cols = profile.header_row, dict(profile.columns)
    else:
        header_row, cols = find_header(ws)

    start = header_row + 1 + (profile.skip_rows_after_header if profile else 0)

    lines: list[InvoiceLine] = []
    declared_amount: Decimal | None = None   # итог без НДС
    declared_total: Decimal | None = None    # итог с НДС

    def cell(row_vals, fld):
        idx = cols.get(fld)
        if idx is None or idx >= len(row_vals):
            return None
        return row_vals[idx]

    for excel_row, row_vals in enumerate(
        ws.iter_rows(min_row=start, values_only=True), start=start
    ):
        if row_vals is None or all(v is None for v in row_vals):
            continue

        name = str(cell(row_vals, "name") or "").strip()

        # Итоговая строка: забираем суммы и прекращаем разбор позиций.
        # Критично брать итог ИЗ ТОЙ ЖЕ колонки, по которой считаем строки, иначе
        # сравним сумму без НДС с итогом с НДС и получим ложное расхождение на размер НДС.
        if name and STOP_ROW_RE.match(name):
            declared_amount = to_decimal(cell(row_vals, "amount")) or declared_amount
            declared_total = to_decimal(cell(row_vals, "total")) or declared_total
            break

        qty = to_decimal(cell(row_vals, "qty"))
        if not name or qty is None:
            continue
        if qty <= 0:
            warnings.append(f"Строка {excel_row}: количество {qty} — пропущена.")
            continue

        unit, unit_raw = canon_unit(cell(row_vals, "unit"))
        if unit_raw and not unit:
            warnings.append(f"Строка {excel_row}: неизвестная единица «{unit_raw}».")

        art_val = cell(row_vals, "article")
        article = str(art_val).strip() if art_val not in (None, "") else None

        lines.append(InvoiceLine(
            row=excel_row,
            name=name,
            qty=qty,
            unit=unit,
            unit_raw=unit_raw,
            article=article,
            price=to_decimal(cell(row_vals, "price")),
            amount=to_decimal(cell(row_vals, "amount")),
            vat_rate=(str(cell(row_vals, "vat_rate")).strip()
                      if cell(row_vals, "vat_rate") not in (None, "") else None),
        ))

    if not lines:
        raise ParseError(
            f"Заголовки найдены (строка {header_row}), но ни одной позиции не разобрано. "
            "Похоже, раскладка колонок не совпала — задайте профиль поставщика."
        )

    result = ParseResult(
        lines=lines,
        file_hash=fh,
        header_row=header_row,
        columns=cols,
        sheet=ws.title,
        declared_amount=declared_amount,
        declared_total=declared_total,
        warnings=warnings,
    )

    # Контрольная сумма. Сверяем только однородные величины: сумма строк (без НДС)
    # против итога без НДС. Допуск — по копейке на строку (округление у поставщика).
    if declared_amount is not None:
        diff = (result.lines_total - declared_amount).copy_abs()
        tolerance = Decimal("0.01") * len(lines)
        if diff > tolerance:
            result.warnings.append(
                f"Сумма строк {result.lines_total} не сходится с итогом в файле "
                f"{declared_amount} (расхождение {diff}). Документ — на проверку."
            )
    else:
        result.warnings.append("Итог без НДС в файле не найден — сверить сумму вручную.")

    return result


def profile_from_result(res: ParseResult, inn: str | None = None) -> SupplierProfile:
    """Из удачного разбора собрать профиль, чтобы сохранить и переиспользовать."""
    return SupplierProfile(
        inn=inn,
        sheet=res.sheet,
        header_row=res.header_row,
        columns=dict(res.columns),
    )
